import openpyxl
from openpyxl import Workbook
import os
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
import smtplib  # For sending emails
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Sample data for student information
students_data = {
    "AIML": [
        {"Enrollment": "00101192023", "Name": "Vibhuti", "Password": "pass123", "Email": "vibhuti076btaiml23@igdtuw.ac.in"},
        {"Enrollment": "05301192023", "Name": "Rupali", "Password": "pass123", "Email": "rupali053btaiml23@igdtuw.ac.in"}
    ],
    "CSE": [
        {"Enrollment": "05401192023", "Name": "Saisha", "Password": "pass123", "Email": "saisha054btaiml23@igdtuw.ac.in"},
        {"Enrollment": "04101192023", "Name": "Nandini", "Password": "pass123", "Email": "nandini041btaiml23@igdtuw.ac.in"}
    ]
}

# Email configuration
SENDER_EMAIL = "vibhutigoyal0603@gmail.com"
SENDER_PASSWORD = "**************"
SMTP_SERVER = "smtp.gmail.com"  # e.g., "smtp.gmail.com"
SMTP_PORT = 587  # Typical SMTP port

# Initialize the SQLite database
def initialize_db():
    connection = sqlite3.connect("attendance.db")
    cursor = connection.cursor()
    
    # Drop and create the attendance table
    cursor.execute("DROP TABLE IF EXISTS attendance")
    cursor.execute("""
    CREATE TABLE attendance (
        date TEXT,
        branch TEXT,
        enrollment TEXT,
        status TEXT)""")
    
    # Create the students table with credentials
    cursor.execute("DROP TABLE IF EXISTS students")
    cursor.execute("""
    CREATE TABLE students (
        enrollment TEXT PRIMARY KEY,
        name TEXT,
        email TEXT,
        password TEXT,
        branch TEXT)""")
    
    # Add sample student records
    students_list = [
        ("00101192023", "Vibhuti", "vibhuti076btaiml23@igdtuw.ac.in","password123", "AIML"),
        ("05301192023", "Rupali", "rupali053btaiml23@igdtuw.ac.in","password123", "AIML"),
        ("05401192023", "Saisha", "saisha054btaiml23@igdtuw.ac.in","password123", "CSE"),
        ("04101192023", "Nandini", "nandini041btaiml23@igdtuw.ac.in","password123", "CSE"),
    ]
    cursor.executemany("INSERT OR IGNORE INTO students VALUES (?, ?, ?, ?, ?)", students_list)

    connection.commit()
    connection.close()

initialize_db()


# Main application class
class AttendanceApp:
    def _init_(self, root):
        self.root = root
        self.root.title("Attendance Management System")
        self.root.geometry("800x500")

        self.create_login_screen()


    def create_login_screen(self):
        """Create login screen with options to log in as Student or Teacher."""
        self.clear_screen()

        tk.Label(self.root, text="Login Page", font=("Arial", 18, "bold")).pack(pady=20)

        tk.Button(self.root, text="Login as Teacher", command=self.teacher_login).pack(pady=10)
        tk.Button(self.root, text="Login as Student", command=self.student_login).pack(pady=10)

    def teacher_login(self):
        """Display login fields for the teacher."""
        self.clear_screen()
        tk.Label(self.root, text="Teacher Login", font=("Arial", 18, "bold")).pack(pady=20)

        tk.Label(self.root, text="Teacher ID").pack()
        self.teacher_id_entry = tk.Entry(self.root)
        self.teacher_id_entry.pack(pady=5)

        tk.Label(self.root, text="Password").pack()
        self.teacher_password_entry = tk.Entry(self.root, show="*")
        self.teacher_password_entry.pack(pady=5)

        tk.Button(self.root, text="Login", command=self.verify_teacher_login).pack(pady=20)

    def verify_teacher_login(self):
        """Verify teacher credentials (placeholder for actual authentication)."""
        teacher_id = self.teacher_id_entry.get()
        password = self.teacher_password_entry.get()
        
        # Replace with actual credential verification
        if teacher_id == "teacher" and password == "password":
            self.show_teacher_dashboard()
        else:
            messagebox.showerror("Error", "Invalid credentials")

    # Add this method to the AttendanceApp class
    def show_teacher_dashboard(self):
        self.clear_screen()
        tk.Label(self.root, text="Teacher Dashboard", font=("Arial", 18, "bold")).pack(pady=20)

        tk.Label(self.root, text="Select Date").pack()
        
        # Calendar date entry
        self.date_entry = DateEntry(self.root, width=12, background="darkblue", foreground="white", date_pattern='yyyy-mm-dd')
        self.date_entry.pack(pady=5)

        tk.Label(self.root, text="Select Branch").pack()
        self.branch_var = tk.StringVar()
        branch_menu = ttk.Combobox(self.root, textvariable=self.branch_var, values=list(students_data.keys()))
        branch_menu.pack(pady=5)
        branch_menu.bind("<<ComboboxSelected>>", self.load_student_list)

        self.table_frame = tk.Frame(self.root)
        self.table_frame.pack(fill="both", expand=True, pady=10)

        tk.Button(self.root, text="Save Attendance", command=self.save_attendance).pack(pady=10)
        tk.Button(self.root, text="Analyze Attendance", command=self.analyze_attendance).pack(pady=5)
        
        # Button to view the Excel sheet
        tk.Button(self.root, text="View Excel Sheet", command=self.open_excel_sheet).pack(pady=5)
        
        
        # Button to view monthly attendance
        tk.Button(self.root, text="View Monthly Attendance", command=self.view_monthly_attendance).pack(pady=5)
    def view_monthly_attendance(self):
        branch = self.branch_var.get()
        if not branch:
            messagebox.showwarning("Warning", "Please select a branch.")
            return

        # Fetch attendance data from the Excel sheet
        file_name = "attendance_records.xlsx"
        if not os.path.exists(file_name):
            messagebox.showinfo("Info", "No attendance records found. Save attendance first.")
            return

        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["Attendance"]

        # Dictionary to hold monthly attendance data
        monthly_attendance = {}

        # Read data from Excel sheet to count attendance by month and student
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            date, branch_name, enrollment, name, status = row
            if branch_name == branch:
                month_year = date[:7]  # Extract the YYYY-MM part of the date
                if month_year not in monthly_attendance:
                    monthly_attendance[month_year] = {}
                if enrollment not in monthly_attendance[month_year]:
                    monthly_attendance[month_year][enrollment] = {"name": name, "present_days": 0, "total_days": 0}
                monthly_attendance[month_year][enrollment]["total_days"] += 1
                if status == "Present":
                    monthly_attendance[month_year][enrollment]["present_days"] += 1

        # Create a new window to show the monthly attendance summary
        monthly_window = tk.Toplevel(self.root)
        monthly_window.title("Monthly Attendance Summary")
        monthly_window.geometry("600x400")

        tk.Label(monthly_window, text=f"Monthly Attendance for {branch} Branch", font=("Arial", 14, "bold")).pack(pady=10)

        # Iterate through the months and display attendance for each student
        for month_year, students in monthly_attendance.items():
            month_label = tk.Label(monthly_window, text=f"Month: {month_year}", font=("Arial", 12, "bold"))
            month_label.pack(pady=5)

            for enrollment, data in students.items():
                name = data["name"]
                total_days = data["total_days"]
                present_days = data["present_days"]
                present_percentage = (present_days / total_days) * 100 if total_days > 0 else 0

                student_label = f"{name} ({enrollment}) - {present_percentage:.2f}% Attendance"
                tk.Label(monthly_window, text=student_label).pack(pady=2)

        workbook.close()

        
    def open_excel_sheet(self):
        """Open the attendance Excel sheet."""
        file_name = "attendance_records.xlsx"
        if os.path.exists(file_name):
            try:
                # Open the file using the default application
                if os.name == 'nt':  # For Windows
                    os.startfile(file_name)
                else:  # For macOS and Linux
                    subprocess.call(['open' if os.name == 'posix' else 'xdg-open', file_name])
            except Exception as e:
                messagebox.showerror("Error", f"Could not open the Excel sheet: {e}")
        else:
            messagebox.showinfo("Info", "No attendance records found. Save attendance first.")

    def load_student_list(self, event):
        """Load the student list for the selected branch."""
        branch = self.branch_var.get()
        students = students_data.get(branch, [])
        
        # Clear existing table if any
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        tk.Label(self.table_frame, text="Enrollment No.", font=("Arial", 10, "bold")).grid(row=0, column=0)
        tk.Label(self.table_frame, text="Student Name", font=("Arial", 10, "bold")).grid(row=0, column=1)
        tk.Label(self.table_frame, text="Email", font=("Arial", 10, "bold")).grid(row=0, column=2)
        tk.Label(self.table_frame, text="Attendance", font=("Arial", 10, "bold")).grid(row=0, column=3)

        self.attendance_vars = {}
        for i, student in enumerate(students, start=1):
            enrollment = student["Enrollment"]
            name = student["Name"]
            email = student["Email"]

            tk.Label(self.table_frame, text=enrollment).grid(row=i, column=0)
            tk.Label(self.table_frame, text=name).grid(row=i, column=1)
            tk.Label(self.table_frame, text=email).grid(row=i, column=2)

            attendance_var = tk.StringVar(value="Absent")
            tk.Radiobutton(self.table_frame, text="Present", variable=attendance_var, value="Present").grid(row=i, column=3, sticky="w")
            tk.Radiobutton(self.table_frame, text="Absent", variable=attendance_var, value="Absent").grid(row=i, column=4, sticky="e")
            self.attendance_vars[enrollment] = attendance_var

    def save_attendance(self):
        """Save the attendance data to the database and an Excel sheet."""
        date = self.date_entry.get()
        branch = self.branch_var.get()

        # Saving to SQLite database
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()

        for enrollment, attendance_var in self.attendance_vars.items():
            status = attendance_var.get()
            cursor.execute("INSERT INTO attendance (date, branch, enrollment, status) VALUES (?, ?, ?, ?)",
                           (date, branch, enrollment, status))

        conn.commit()
        conn.close()

        # Saving to Excel sheet
        file_name = "attendance_records.xlsx"
        if not os.path.exists(file_name):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Attendance"
            # Add headers to the Excel sheet
            sheet.append(["Date", "Branch", "Enrollment No.", "Student Name", "Status"])
        else:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook["Attendance"]

        # Add attendance records to Excel
        for enrollment, attendance_var in self.attendance_vars.items():
            status = attendance_var.get()
            # Find student name based on enrollment number
            student_name = next((student["Name"] for student in students_data[branch] if student["Enrollment"] == enrollment), "Unknown")
            sheet.append([date, branch, enrollment, student_name, status])

        workbook.save(file_name)

        messagebox.showinfo("Success", "Attendance saved successfully in the database and Excel sheet!")

    def send_email(self, recipient_email, student_name, present_percentage):
        """Send an email notification to a student with low attendance."""
        message = MIMEMultipart()
        message["From"] = SENDER_EMAIL
        message["To"] = recipient_email
        message["Subject"] = "Low Attendance Alert"

        body = (f"Dear {student_name},\n\n"
                f"We noticed that your attendance is below the required threshold. Your current attendance rate is {present_percentage:.2f}%.\n"
                "Please reach out to your teacher if you need assistance in improving your attendance.\n\n"
                "Best regards,\nIGDTUW")
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.sendmail(SENDER_EMAIL, recipient_email, message.as_string())
            messagebox.showinfo("Success", f"Email sent to {student_name}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not send email to {student_name}: {e}")
    def analyze_attendance(self):
        file_name = "attendance_records.xlsx"
        if not os.path.exists(file_name):
            messagebox.showinfo("Info", "No attendance records found. Save attendance first.")
            return

        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["Attendance"]

        # Dictionary to hold attendance data for each student
        student_attendance = {}

        # Read data from Excel sheet to count total and present days for each student
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            date, branch, enrollment, name, status = row
            if enrollment not in student_attendance:
                student_attendance[enrollment] = {"name": name, "present_days": 0, "total_days": 0}
            student_attendance[enrollment]["total_days"] += 1
            if status == "Present":
                student_attendance[enrollment]["present_days"] += 1

        # Create a new window for the analysis result
        analysis_window = tk.Toplevel(self.root)
        analysis_window.title("Attendance Analysis")
        analysis_window.geometry("500x400")

        tk.Label(analysis_window, text="Students with Low Attendance (<75%)", font=("Arial", 12, "bold")).pack(pady=10)

        # Iterate through the attendance data to find low attendance students
        for enrollment, data in student_attendance.items():
            total_days = data["total_days"]
            present_days = data["present_days"]
            name = data["name"]

            if total_days > 0:
                present_percentage = (present_days / total_days) * 100
                if present_percentage < 75:  # Check if attendance is below 75%
                    label_text = f"{name} ({enrollment}) - {present_percentage:.2f}%"
                    student_frame = tk.Frame(analysis_window)
                    student_frame.pack(fill="x", padx=10, pady=2)

                    tk.Label(student_frame, text=label_text, fg="red").pack(side="left")

                    # Find the student's email using the enrollment number from all branches
                    student_email = None
                    for branch_students in students_data.values():
                        student = next((s for s in branch_students if s["Enrollment"] == enrollment), None)
                        if student:
                            student_email = student["Email"]
                            break

                    # Only create the email button if the email was found
                    if student_email:
                        email_button = tk.Button(
                            student_frame, 
                            text="Email", 
                            command=lambda e=student_email, n=name, p=present_percentage: self.send_email(e, n, p)
                        )
                        email_button.pack(side="right")

        workbook.close()

    def student_login(self):
        self.clear_screen()

        tk.Label(self.root, text="Student Login", font=("Arial", 18, "bold")).pack(pady=20)

        tk.Label(self.root, text="Enrollment Number").pack()
        self.student_enroll_entry = tk.Entry(self.root)
        self.student_enroll_entry.pack(pady=5)

        tk.Label(self.root, text="Password").pack()
        self.student_password_entry = tk.Entry(self.root, show="*")
        self.student_password_entry.pack(pady=5)

        tk.Button(self.root, text="Login", command=self.verify_student_login).pack(pady=20)

    def verify_student_login(self):
        enroll_number = self.student_enroll_entry.get()
        password = self.student_password_entry.get()

        for branch, students in students_data.items():
            for student in students:
                if student["Enrollment"] == enroll_number and student["Password"] == password:
                    self.current_student = student
                    self.show_student_dashboard()
                    return

        messagebox.showerror("Error", "Invalid enrollment number or password")




    def show_student_dashboard(self):
        self.clear_screen()
        tk.Label(self.root, text="Student Dashboard", font=("Arial", 18, "bold")).pack(pady=20)

        tk.Button(self.root, text="View Per Day Attendance", command=self.view_per_day_attendance).pack(pady=10)
        tk.Button(self.root, text="View Overall Attendance Percentage", command=self.view_overall_percentage).pack(pady=10)

    def view_per_day_attendance(self):
        file_name = "attendance_records.xlsx"
        if not os.path.exists(file_name):
            messagebox.showinfo("Info", "No attendance records found.")
            return

        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["Attendance"]

        student_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date, branch, enrollment, name, status = row
            if enrollment == self.current_student["Enrollment"]:
                student_data.append((date, status))

        workbook.close()

        attendance_window = tk.Toplevel(self.root)
        attendance_window.title("Per Day Attendance")
        tk.Label(attendance_window, text="Date - Status", font=("Arial", 14, "bold")).pack(pady=10)

        for date, status in student_data:
            tk.Label(attendance_window, text=f"{date} - {status}").pack()

    def view_overall_percentage(self):
        file_name = "attendance_records.xlsx"
        if not os.path.exists(file_name):
            messagebox.showinfo("Info", "No attendance records found.")
            return

        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook["Attendance"]

        total_days = 0
        present_days = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date, branch, enrollment, name, status = row
            if enrollment == self.current_student["Enrollment"]:
                total_days += 1
                if status == "Present":
                    present_days += 1

        workbook.close()

        overall_percentage = (present_days / total_days * 100) if total_days > 0 else 0
        messagebox.showinfo("Overall Attendance", f"Your overall attendance is {overall_percentage:.2f}%")
    
            

    def clear_screen(self):
        """Clear all widgets from the root window."""
        for widget in self.root.winfo_children():
            widget.destroy()

# Initialize the Tkinter app
root = tk.Tk()
app = AttendanceApp(root)
root.mainloop()
